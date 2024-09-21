#needs to insert data into the desired sheet
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

workbook = load_workbook("/Users/ryantiller/personal-finance-project/Finances-master.xlsx")

def data_insert_monthly_spend(data,month_year):
    totals = {
    'grocery' : 0,
    'restaurants' : 0,
    'venmo': 0,
    'coffee': 0,
    'gas' : 0,
    'car' : 0,
    'rent' : 0,
    'trasportation' : 0, 
    'misc': 0,
    'poker' : 0,
    'Amazon' : 0,
    'Total Out': 0,
    }

    new_month = workbook.create_sheet(title=f"{month_year}")  
    new_month.cell(row=1, column=1,value='expenses') 
    for col_index, (category, prices) in enumerate(data.items(), start=2):
        new_month.cell(row=1,column=col_index, value=category)
        for i in range(0,len(prices)):
            stripped = prices[i].replace('$', '')
            stripped_again = stripped.replace(',', '')
            floated = float(stripped_again)
            totals[category] = totals[category] + floated
            totals['Total Out'] = totals['Total Out'] + floated
            new_month.cell(row=2+i,column=col_index,value=floated)

    # create table for totals
    new_month.cell(row=1, column=17,value='Category Totals')
    for col_index_2, (category, sum) in enumerate(totals.items(), start=17):
        new_month.cell(row=2,column=col_index_2, value=category)
        new_month.cell(row=3,column=col_index_2, value=sum)
        
        
    # this goes last
    workbook.save("/Users/ryantiller/personal-finance-project/Finances-master.xlsx")




